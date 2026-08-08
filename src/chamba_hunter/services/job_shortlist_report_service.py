from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from chamba_hunter.repositories.job_shortlist_report_repository import (
    JobShortlistReportRepository,
    ShortlistReportRow,
    ShortlistReportSource,
)
from chamba_hunter.services.job_operational_priority_service import (
    ACTIVE_STATES,
    CHANNEL_RANK,
    MATCH_LEVEL_RANK,
    STATE_RANK,
)


REPORT_VERSION = "SHORTLIST_REPORT_V1"
DEFAULT_PROFILE_NAME = "BACKEND_SOFTWARE_V1"


_LEVEL_FILLS = {
    "VERY_HIGH": "C6EFCE",
    "HIGH": "E2F0D9",
    "MEDIUM": "FFF2CC",
    "LOW": "E7E6E6",
}


_STATE_FILLS = {
    "NEW": "DDEBF7",
    "UPDATED": "FCE4D6",
    "KNOWN": "FFFFFF",
    "INACTIVE": "E7E6E6",
    "SUPERSEDED": "E7E6E6",
    "OUT_OF_SCOPE": "E7E6E6",
}


_TITLE_FILL = "1F4E78"
_HEADER_FILL = "D9EAF7"
_SUBTITLE_FILL = "EAF2F8"
_BORDER_COLOR = "D9E1F2"
_LINK_COLOR = "0563C1"


_HEADERS = (
    "Priority Rank",
    "Operational State",
    "Match Level",
    "Professional Score",
    "Company",
    "Title",
    "Origin / Provider",
    "Application Channel",
    "Open",
    "Tracked Status",
    "First Seen",
    "Last Changed",
    "Published At",
    "Same-title Count",
    "Occupation",
    "Backend Relevance",
    "Seniority",
    "Leadership",
    "Role Pts",
    "Skills Pts",
    "Seniority Pts",
    "Leadership Pts",
    "Tech Penalty",
    "Score Ceiling",
    "Exact Skills",
    "Peer Skills",
    "Related Skills",
    "Secondary Skills",
    "Alternate Stack",
    "Ceiling Reasons",
    "Application Type",
    "Applied At",
    "Record Kind",
    "Record ID",
    "Application Target",
    "Job URL",
    "Apply URL",
)


_WIDTHS = {
    1: 12,
    2: 16,
    3: 13,
    4: 12,
    5: 24,
    6: 44,
    7: 18,
    8: 24,
    9: 11,
    10: 16,
    11: 19,
    12: 19,
    13: 19,
    14: 14,
    15: 22,
    16: 18,
    17: 14,
    18: 14,
    19: 11,
    20: 11,
    21: 13,
    22: 14,
    23: 12,
    24: 12,
    25: 34,
    26: 28,
    27: 34,
    28: 28,
    29: 24,
    30: 34,
    31: 18,
    32: 19,
    33: 12,
    34: 10,
    35: 42,
    36: 42,
    37: 42,
}


@dataclass(frozen=True, slots=True)
class ShortlistReportItem:
    source: ShortlistReportRow

    occupation_class: str | None
    backend_relevance: str | None
    seniority_class: str | None
    leadership_class: str | None

    exact_skills: tuple[str, ...]
    peer_skills: tuple[str, ...]
    related_skills: tuple[str, ...]
    secondary_skills: tuple[str, ...]
    alternate_families: tuple[str, ...]
    ceiling_reasons: tuple[str, ...]

    normalized_title: str
    same_title_count: int = 1

    @property
    def actionable(
        self,
    ) -> bool:
        return (
            self.source.operational_state
            in ACTIVE_STATES
        )


@dataclass(frozen=True, slots=True)
class ShortlistReportSummary:
    source: ShortlistReportSource

    focus: tuple[
        ShortlistReportItem,
        ...
    ]

    high_value: tuple[
        ShortlistReportItem,
        ...
    ]

    all_current: tuple[
        ShortlistReportItem,
        ...
    ]

    history: tuple[
        ShortlistReportItem,
        ...
    ]

    state_counts: dict[str, int]
    level_counts: dict[str, int]
    channel_counts: dict[str, int]

    generated_at: datetime


def _json_object(
    raw: str | None,
) -> dict:
    if not raw:
        return {}

    parsed = json.loads(
        raw
    )

    if not isinstance(
        parsed,
        dict,
    ):
        return {}

    return parsed


def _string_tuple(
    value,
) -> tuple[str, ...]:
    if not isinstance(
        value,
        list,
    ):
        return ()

    return tuple(
        str(item)
        for item in value
    )


def _normalize_title(
    value: str,
) -> str:
    text = unicodedata.normalize(
        "NFKD",
        value,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(
            character
        )
    )

    text = text.lower()

    text = re.sub(
        r"[^a-z0-9+#.]+",
        " ",
        text,
    )

    return " ".join(
        text.split()
    )


def _build_item(
    row: ShortlistReportRow,
) -> ShortlistReportItem:
    professional = _json_object(
        row.professional_reasons_json
    )

    candidate = professional.get(
        "candidate",
        {},
    )

    if not isinstance(
        candidate,
        dict,
    ):
        candidate = {}

    components = professional.get(
        "components",
        {},
    )

    if not isinstance(
        components,
        dict,
    ):
        components = {}

    skills = components.get(
        "skills",
        {},
    )

    if not isinstance(
        skills,
        dict,
    ):
        skills = {}

    technology_penalty = components.get(
        "technology_penalty",
        {},
    )

    if not isinstance(
        technology_penalty,
        dict,
    ):
        technology_penalty = {}

    return ShortlistReportItem(
        source=row,
        occupation_class=(
            str(
                candidate[
                    "occupation_class"
                ]
            )
            if candidate.get(
                "occupation_class"
            )
            is not None
            else None
        ),
        backend_relevance=(
            str(
                candidate[
                    "backend_relevance"
                ]
            )
            if candidate.get(
                "backend_relevance"
            )
            is not None
            else None
        ),
        seniority_class=(
            str(
                candidate[
                    "seniority_class"
                ]
            )
            if candidate.get(
                "seniority_class"
            )
            is not None
            else None
        ),
        leadership_class=(
            str(
                candidate[
                    "leadership_class"
                ]
            )
            if candidate.get(
                "leadership_class"
            )
            is not None
            else None
        ),
        exact_skills=_string_tuple(
            skills.get(
                "exact"
            )
        ),
        peer_skills=_string_tuple(
            skills.get(
                "peer"
            )
        ),
        related_skills=_string_tuple(
            skills.get(
                "related"
            )
        ),
        secondary_skills=_string_tuple(
            skills.get(
                "secondary"
            )
        ),
        alternate_families=(
            _string_tuple(
                technology_penalty.get(
                    "alternate_families"
                )
            )
        ),
        ceiling_reasons=_string_tuple(
            professional.get(
                "ceiling_reasons"
            )
        ),
        normalized_title=(
            _normalize_title(
                row.title
            )
        ),
    )


def _sort_key(
    item: ShortlistReportItem,
) -> tuple:
    source = item.source

    return (
        -int(
            item.actionable
        ),
        -MATCH_LEVEL_RANK.get(
            source.professional_match_level,
            0,
        ),
        -STATE_RANK.get(
            source.operational_state,
            0,
        ),
        -source.professional_score,
        -CHANNEL_RANK.get(
            source.application_channel,
            0,
        ),
        -source.first_seen_at.timestamp(),
        source.company_name.lower(),
        source.title.lower(),
        source.record_kind,
        source.record_id,
    )


def _with_duplicate_counts(
    items: list[
        ShortlistReportItem
    ],
) -> list[
    ShortlistReportItem
]:
    counts = Counter(
        (
            item.source.company_name.casefold(),
            item.normalized_title,
        )
        for item in items
        if item.actionable
    )

    return [
        ShortlistReportItem(
            source=item.source,
            occupation_class=(
                item.occupation_class
            ),
            backend_relevance=(
                item.backend_relevance
            ),
            seniority_class=(
                item.seniority_class
            ),
            leadership_class=(
                item.leadership_class
            ),
            exact_skills=(
                item.exact_skills
            ),
            peer_skills=(
                item.peer_skills
            ),
            related_skills=(
                item.related_skills
            ),
            secondary_skills=(
                item.secondary_skills
            ),
            alternate_families=(
                item.alternate_families
            ),
            ceiling_reasons=(
                item.ceiling_reasons
            ),
            normalized_title=(
                item.normalized_title
            ),
            same_title_count=(
                counts.get(
                    (
                        item.source
                        .company_name
                        .casefold(),
                        item.normalized_title,
                    ),
                    1,
                )
                if item.actionable
                else 1
            ),
        )
        for item in items
    ]


def build_summary(
    source: ShortlistReportSource,
) -> ShortlistReportSummary:
    items = [
        _build_item(
            row
        )
        for row in source.rows
    ]

    items = _with_duplicate_counts(
        items
    )

    ranked = sorted(
        items,
        key=_sort_key,
    )

    current = tuple(
        item
        for item in ranked
        if item.actionable
    )

    focus = tuple(
        item
        for item in current
        if (
            item.source
            .operational_state
            in {
                "NEW",
                "UPDATED",
            }
            and item.source
            .professional_match_level
            in {
                "VERY_HIGH",
                "HIGH",
            }
        )
    )

    high_value = tuple(
        item
        for item in current
        if item.source
        .professional_match_level
        in {
            "VERY_HIGH",
            "HIGH",
        }
    )

    history = tuple(
        item
        for item in ranked
        if not item.actionable
    )

    state_counts = Counter(
        item.source
        .operational_state
        for item in ranked
    )

    level_counts = Counter(
        item.source
        .professional_match_level
        for item in ranked
    )

    channel_counts = Counter(
        item.source
        .application_channel
        for item in ranked
    )

    return ShortlistReportSummary(
        source=source,
        focus=focus,
        high_value=high_value,
        all_current=current,
        history=history,
        state_counts=dict(
            state_counts
        ),
        level_counts=dict(
            level_counts
        ),
        channel_counts=dict(
            channel_counts
        ),
        generated_at=datetime.now(
            timezone.utc
        ),
    )


def _excel_datetime(
    value: datetime | None,
) -> datetime | None:
    if value is None:
        return None

    if value.tzinfo is None:
        return value

    return (
        value
        .astimezone(
            timezone.utc
        )
        .replace(
            tzinfo=None
        )
    )


def _join(
    values: tuple[str, ...],
) -> str:
    return ", ".join(
        values
    )


def _item_values(
    rank: int,
    item: ShortlistReportItem,
) -> tuple:
    source = item.source

    return (
        rank,
        source.operational_state,
        source.professional_match_level,
        source.professional_score,
        source.company_name,
        source.title,
        source.origin,
        source.application_channel,
        (
            "Open"
            if source.application_target
            else ""
        ),
        source.application_status,
        _excel_datetime(
            source.first_seen_at
        ),
        _excel_datetime(
            source.last_changed_at
        ),
        _excel_datetime(
            source.published_at
        ),
        item.same_title_count,
        item.occupation_class,
        item.backend_relevance,
        item.seniority_class,
        item.leadership_class,
        source.role_score,
        source.skills_score,
        source.seniority_score,
        source.leadership_score,
        source.technology_penalty,
        source.score_ceiling,
        _join(
            item.exact_skills
        ),
        _join(
            item.peer_skills
        ),
        _join(
            item.related_skills
        ),
        _join(
            item.secondary_skills
        ),
        _join(
            item.alternate_families
        ),
        _join(
            item.ceiling_reasons
        ),
        source.application_type,
        _excel_datetime(
            source.application_applied_at
        ),
        source.record_kind,
        source.record_id,
        source.application_target,
        source.job_url,
        source.apply_url,
    )


def _style_link(
    cell: Cell,
    target: str | None,
) -> None:
    if not target:
        return

    cell.hyperlink = target
    cell.font = Font(
        color=_LINK_COLOR,
        underline="single",
    )


def _create_overview(
    workbook: Workbook,
    summary: ShortlistReportSummary,
) -> None:
    sheet = workbook.active
    sheet.title = "Overview"
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells(
        "A1:F1"
    )

    title = sheet["A1"]
    title.value = "Chamba Hunter — Shortlist"
    title.fill = PatternFill(
        "solid",
        fgColor=_TITLE_FILL,
    )
    title.font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )
    title.alignment = Alignment(
        vertical="center",
    )
    sheet.row_dimensions[1].height = 28

    metadata = (
        ("Report version", REPORT_VERSION),
        (
            "Search profile",
            summary.source.profile_name,
        ),
        (
            "Priority run",
            summary.source.priority_run_id,
        ),
        (
            "Priority run finished (UTC)",
            _excel_datetime(
                summary.source
                .priority_run_finished_at
            ),
        ),
        (
            "Generated (UTC)",
            _excel_datetime(
                summary.generated_at
            ),
        ),
        (
            "Priority rule",
            (
                summary.source.rows[0]
                .priority_rule_version
                if summary.source.rows
                else ""
            ),
        ),
        (
            "Professional rule",
            (
                summary.source.rows[0]
                .professional_rule_version
                if summary.source.rows
                else ""
            ),
        ),
    )

    sheet["A3"] = "Metadata"
    sheet["A3"].font = Font(
        bold=True,
        color="FFFFFF",
    )
    sheet["A3"].fill = PatternFill(
        "solid",
        fgColor=_TITLE_FILL,
    )

    for row_index, (
        label,
        value,
    ) in enumerate(
        metadata,
        start=4,
    ):
        sheet.cell(
            row=row_index,
            column=1,
            value=label,
        ).font = Font(
            bold=True
        )

        cell = sheet.cell(
            row=row_index,
            column=2,
            value=value,
        )

        if isinstance(
            value,
            datetime,
        ):
            cell.number_format = (
                "yyyy-mm-dd hh:mm"
            )

    start = 13

    sheet.cell(
        row=start,
        column=1,
        value="Current counts",
    ).font = Font(
        bold=True,
        color="FFFFFF",
    )

    sheet.cell(
        row=start,
        column=1,
    ).fill = PatternFill(
        "solid",
        fgColor=_TITLE_FILL,
    )

    rows = (
        (
            "Focus",
            len(
                summary.focus
            ),
        ),
        (
            "VERY_HIGH + HIGH",
            len(
                summary.high_value
            ),
        ),
        (
            "All current",
            len(
                summary.all_current
            ),
        ),
        (
            "History",
            len(
                summary.history
            ),
        ),
        (
            "NEW",
            summary.state_counts.get(
                "NEW",
                0,
            ),
        ),
        (
            "UPDATED",
            summary.state_counts.get(
                "UPDATED",
                0,
            ),
        ),
        (
            "KNOWN",
            summary.state_counts.get(
                "KNOWN",
                0,
            ),
        ),
        (
            "VERY_HIGH",
            summary.level_counts.get(
                "VERY_HIGH",
                0,
            ),
        ),
        (
            "HIGH",
            summary.level_counts.get(
                "HIGH",
                0,
            ),
        ),
        (
            "DIRECT_APPLY_URL",
            summary.channel_counts.get(
                "DIRECT_APPLY_URL",
                0,
            ),
        ),
        (
            "JOB_URL",
            summary.channel_counts.get(
                "JOB_URL",
                0,
            ),
        ),
    )

    for offset, (
        label,
        value,
    ) in enumerate(
        rows,
        start=1,
    ):
        row_index = start + offset

        sheet.cell(
            row=row_index,
            column=1,
            value=label,
        )

        sheet.cell(
            row=row_index,
            column=2,
            value=value,
        )

    links_start = 13

    sheet.cell(
        row=links_start,
        column=4,
        value="Views",
    ).font = Font(
        bold=True,
        color="FFFFFF",
    )

    sheet.cell(
        row=links_start,
        column=4,
    ).fill = PatternFill(
        "solid",
        fgColor=_TITLE_FILL,
    )

    view_rows = (
        (
            "Focus",
            "NEW/UPDATED + VERY_HIGH/HIGH",
        ),
        (
            "High Value",
            "All current VERY_HIGH/HIGH",
        ),
        (
            "All Current",
            "All actionable rows",
        ),
        (
            "History",
            "Inactive/superseded/out-of-scope",
        ),
    )

    for offset, (
        sheet_name,
        description,
    ) in enumerate(
        view_rows,
        start=1,
    ):
        row_index = (
            links_start
            + offset
        )

        link_cell = sheet.cell(
            row=row_index,
            column=4,
            value=sheet_name,
        )

        link_cell.hyperlink = (
            f"#'{sheet_name}'!A1"
        )

        link_cell.font = Font(
            color=_LINK_COLOR,
            underline="single",
        )

        sheet.cell(
            row=row_index,
            column=5,
            value=description,
        )

    sheet["D20"] = "Notes"
    sheet["D20"].font = Font(
        bold=True,
        color="FFFFFF",
    )
    sheet["D20"].fill = PatternFill(
        "solid",
        fgColor=_TITLE_FILL,
    )

    sheet["D21"] = (
        "Report generation is read-only. "
        "It does not recalculate matching, "
        "freshness, priority, or applications."
    )

    sheet["D22"] = (
        "Same-title Count is informational only; "
        "the report never deduplicates postings."
    )

    sheet["D23"] = (
        "Application status is read from the latest "
        "tracked ATS application when present. "
        "LEAD tracking remains blank until its schema "
        "is explicitly extended."
    )

    for column, width in {
        "A": 28,
        "B": 30,
        "C": 4,
        "D": 24,
        "E": 58,
        "F": 20,
    }.items():
        sheet.column_dimensions[
            column
        ].width = width

    for row in range(
        1,
        25,
    ):
        sheet.row_dimensions[
            row
        ].height = 20

    sheet["D21"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )
    sheet["D22"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )
    sheet["D23"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    sheet.freeze_panes = "A3"


def _style_data_sheet(
    sheet,
    *,
    title: str,
    subtitle: str,
) -> None:
    sheet.sheet_view.showGridLines = False

    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(
            _HEADERS
        ),
    )

    title_cell = sheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.fill = PatternFill(
        "solid",
        fgColor=_TITLE_FILL,
    )

    title_cell.font = Font(
        color="FFFFFF",
        bold=True,
        size=15,
    )

    title_cell.alignment = Alignment(
        vertical="center",
    )

    sheet.row_dimensions[1].height = 27

    sheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=len(
            _HEADERS
        ),
    )

    subtitle_cell = sheet.cell(
        row=2,
        column=1,
        value=subtitle,
    )

    subtitle_cell.fill = PatternFill(
        "solid",
        fgColor=_SUBTITLE_FILL,
    )

    subtitle_cell.alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    sheet.row_dimensions[2].height = 32

    thin = Side(
        style="thin",
        color=_BORDER_COLOR,
    )

    for column_index, header in enumerate(
        _HEADERS,
        start=1,
    ):
        cell = sheet.cell(
            row=4,
            column=column_index,
            value=header,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=_HEADER_FILL,
        )

        cell.font = Font(
            bold=True,
            color="1F1F1F",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = Border(
            bottom=thin,
        )

    sheet.row_dimensions[4].height = 32
    sheet.freeze_panes = "E5"

    for column_index, width in (
        _WIDTHS.items()
    ):
        letter = sheet.cell(
            row=4,
            column=column_index,
        ).column_letter

        sheet.column_dimensions[
            letter
        ].width = width


def _create_data_sheet(
    workbook: Workbook,
    *,
    name: str,
    title: str,
    subtitle: str,
    items: tuple[
        ShortlistReportItem,
        ...
    ],
    table_name: str,
) -> None:
    sheet = workbook.create_sheet(
        name
    )

    _style_data_sheet(
        sheet,
        title=title,
        subtitle=subtitle,
    )

    for rank, item in enumerate(
        items,
        start=1,
    ):
        row_index = (
            rank
            + 4
        )

        values = _item_values(
            rank,
            item,
        )

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            cell = sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(
                    column_index
                    in {
                        5,
                        6,
                        25,
                        26,
                        27,
                        28,
                        29,
                        30,
                        35,
                        36,
                        37,
                    }
                ),
            )

        source = item.source

        state_cell = sheet.cell(
            row=row_index,
            column=2,
        )

        state_cell.fill = PatternFill(
            "solid",
            fgColor=_STATE_FILLS.get(
                source.operational_state,
                "FFFFFF",
            ),
        )

        level_cell = sheet.cell(
            row=row_index,
            column=3,
        )

        level_cell.fill = PatternFill(
            "solid",
            fgColor=_LEVEL_FILLS.get(
                source
                .professional_match_level,
                "FFFFFF",
            ),
        )

        sheet.cell(
            row=row_index,
            column=4,
        ).number_format = "0.0"

        for column_index in (
            19,
            20,
            21,
            22,
            23,
            24,
        ):
            sheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = "0.0"

        for column_index in (
            11,
            12,
            13,
            32,
        ):
            sheet.cell(
                row=row_index,
                column=column_index,
            ).number_format = (
                "yyyy-mm-dd hh:mm"
            )

        _style_link(
            sheet.cell(
                row=row_index,
                column=9,
            ),
            source.application_target,
        )

        _style_link(
            sheet.cell(
                row=row_index,
                column=35,
            ),
            source.application_target,
        )

        _style_link(
            sheet.cell(
                row=row_index,
                column=36,
            ),
            source.job_url,
        )

        _style_link(
            sheet.cell(
                row=row_index,
                column=37,
            ),
            source.apply_url,
        )

        sheet.row_dimensions[
            row_index
        ].height = 34

    if items:
        last_row = (
            len(items)
            + 4
        )

        table = Table(
            displayName=table_name,
            ref=(
                f"A4:"
                f"{sheet.cell(row=4, column=len(_HEADERS)).column_letter}"
                f"{last_row}"
            ),
        )

        table.tableStyleInfo = (
            TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        )

        sheet.add_table(
            table
        )
    else:
        sheet["A3"] = (
            "No rows for the current "
            "operational snapshot."
        )
        sheet["A3"].font = Font(
            italic=True,
            color="666666",
        )


def export_shortlist(
    *,
    repository: (
        JobShortlistReportRepository
    ),
    output_path: Path,
    profile_name: str = (
        DEFAULT_PROFILE_NAME
    ),
) -> ShortlistReportSummary:
    source = repository.load(
        profile_name
    )

    summary = build_summary(
        source
    )

    workbook = Workbook()

    workbook.properties.title = (
        "Chamba Hunter Shortlist"
    )
    workbook.properties.subject = (
        REPORT_VERSION
    )
    workbook.properties.creator = (
        "Chamba Hunter"
    )
    workbook.properties.description = (
        "Read-only shortlist generated from "
        "persisted operational priority state."
    )

    _create_overview(
        workbook,
        summary,
    )

    _create_data_sheet(
        workbook,
        name="Focus",
        title="Focus — New / Updated High Value",
        subtitle=(
            "Primary queue after refresh: "
            "NEW or UPDATED opportunities that are "
            "VERY_HIGH or HIGH professionally."
        ),
        items=summary.focus,
        table_name="FocusTable",
    )

    _create_data_sheet(
        workbook,
        name="High Value",
        title="High Value — Current VERY_HIGH / HIGH",
        subtitle=(
            "All current actionable VERY_HIGH and HIGH "
            "opportunities, ordered by operational priority."
        ),
        items=summary.high_value,
        table_name="HighValueTable",
    )

    _create_data_sheet(
        workbook,
        name="All Current",
        title="All Current — Actionable Opportunities",
        subtitle=(
            "All current NEW / UPDATED / KNOWN opportunities. "
            "No rows are deduplicated automatically."
        ),
        items=summary.all_current,
        table_name="AllCurrentTable",
    )

    _create_data_sheet(
        workbook,
        name="History",
        title="History — Retained Non-actionable State",
        subtitle=(
            "INACTIVE, SUPERSEDED, and OUT_OF_SCOPE rows "
            "retained by operational priority."
        ),
        items=summary.history,
        table_name="HistoryTable",
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        output_path
    )

    return summary

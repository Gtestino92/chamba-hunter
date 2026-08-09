from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


OPENOFFICE_ACTIONS_VERSION = "OPENOFFICE_ACTIONS_V1"

_DATA_SHEETS = (
    "Focus",
    "High Value",
    "All Current",
)

_HEADER_ROW = 4
_FIRST_DATA_ROW = 5

_APPLY_FILL = "2E7D32"
_APPLY_FONT = "FFFFFF"
_APPLIED_FILL = "C6EFCE"


def _headers(
    sheet,
) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[_HEADER_ROW]
        if cell.value is not None
    }


def _macro_url(
    *,
    record_kind: str,
    record_id: int,
    sheet_index: int,
    row_index: int,
    status_column_index: int,
) -> str:
    return (
        "vnd.sun.star.script:"
        "Standard.ChambaHunterActions.MarkApplied"
        "?language=Basic"
        "&location=application"
        f"&kind={record_kind}"
        f"&id={record_id}"
        f"&sheet={sheet_index}"
        f"&row={row_index}"
        f"&status_col={status_column_index}"
    )


def add_openoffice_application_actions(
    output_path: Path,
) -> int:
    workbook = load_workbook(
        output_path
    )

    action_count = 0

    for sheet_name in _DATA_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[
            sheet_name
        ]
        headers = _headers(
            sheet
        )

        required = {
            "Tracked Status",
            "Record Kind",
            "Record ID",
        }
        missing = (
            required
            - set(headers)
        )

        if missing:
            raise ValueError(
                f"{sheet_name}: missing required "
                "shortlist columns: "
                + ", ".join(
                    sorted(missing)
                )
            )

        status_column = headers[
            "Tracked Status"
        ]
        kind_column = headers[
            "Record Kind"
        ]
        id_column = headers[
            "Record ID"
        ]
        sheet_index = (
            workbook.sheetnames.index(
                sheet_name
            )
        )

        for row in range(
            _FIRST_DATA_ROW,
            sheet.max_row + 1,
        ):
            status_cell = sheet.cell(
                row=row,
                column=status_column,
            )
            status = (
                str(
                    status_cell.value
                ).strip()
                if status_cell.value
                is not None
                else ""
            )

            if status:
                if (
                    status.upper()
                    == "APPLIED"
                ):
                    status_cell.fill = (
                        PatternFill(
                            "solid",
                            fgColor=(
                                _APPLIED_FILL
                            ),
                        )
                    )
                    status_cell.font = Font(
                        bold=True
                    )
                continue

            kind_value = sheet.cell(
                row=row,
                column=kind_column,
            ).value
            id_value = sheet.cell(
                row=row,
                column=id_column,
            ).value

            if (
                kind_value is None
                or id_value is None
            ):
                continue

            record_kind = (
                str(
                    kind_value
                )
                .strip()
                .upper()
            )

            if record_kind not in {
                "ATS",
                "LEAD",
            }:
                continue

            try:
                record_id = int(
                    id_value
                )
            except (
                TypeError,
                ValueError,
            ):
                continue

            status_cell.value = "APPLY"
            status_cell.hyperlink = (
                _macro_url(
                    record_kind=(
                        record_kind
                    ),
                    record_id=record_id,
                    sheet_index=(
                        sheet_index
                    ),
                    row_index=row - 1,
                    status_column_index=(
                        status_column
                        - 1
                    ),
                )
            )
            status_cell.fill = PatternFill(
                "solid",
                fgColor=_APPLY_FILL,
            )
            status_cell.font = Font(
                color=_APPLY_FONT,
                bold=True,
            )
            status_cell.alignment = (
                Alignment(
                    horizontal="center",
                    vertical="center",
                )
            )

            action_count += 1

    workbook.save(
        output_path
    )

    return action_count


def create_openoffice_actions_test(
    output_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test"
    sheet.sheet_view.showGridLines = (
        False
    )

    sheet["A1"] = (
        "Chamba Hunter — "
        "OpenOffice action test"
    )
    sheet["A1"].font = Font(
        bold=True,
        size=14,
        color="FFFFFF",
    )
    sheet["A1"].fill = PatternFill(
        "solid",
        fgColor="1F4E78",
    )

    sheet["A3"] = (
        "Click PING. It must only "
        "show a macro confirmation; "
        "it does not touch the DB."
    )

    cell = sheet["A5"]
    cell.value = "PING"
    cell.hyperlink = (
        "vnd.sun.star.script:"
        "Standard.ChambaHunterActions.Ping"
        "?language=Basic"
        "&location=application"
    )
    cell.fill = PatternFill(
        "solid",
        fgColor=_APPLY_FILL,
    )
    cell.font = Font(
        color=_APPLY_FONT,
        bold=True,
    )
    cell.alignment = Alignment(
        horizontal="center",
    )

    sheet.column_dimensions[
        "A"
    ].width = 72
    sheet.row_dimensions[5].height = 28

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )
    workbook.save(
        output_path
    )

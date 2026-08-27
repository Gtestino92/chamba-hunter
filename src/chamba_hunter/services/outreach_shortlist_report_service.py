from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
)
from openpyxl.utils import (
    get_column_letter,
)

from chamba_hunter.domain.enums import (
    ContactType,
)
from chamba_hunter.repositories.company_outreach_repository import (
    CompanyOutreachRepository,
    OutreachReportRow,
)
from chamba_hunter.services.public_contact_quality import (
    contact_quality_label_for,
    contact_quality_score_for,
)


REPORT_VERSION = (
    "OUTREACH_REPORT_V2"
)

DEFAULT_MIN_SCORE = 45.0
DEFAULT_MIN_EXPLORE_SCORE = 35.0


@dataclass(frozen=True, slots=True)
class OutreachReportSummary:
    priority: int
    explore: int
    history: int


class OutreachShortlistReportService:
    def __init__(
        self,
        repository: CompanyOutreachRepository,
    ) -> None:
        self.repository = repository

    def export(
        self,
        *,
        search_profile_name: str,
        output: Path,
        min_score: float = (
            DEFAULT_MIN_SCORE
        ),
        min_explore_score: float = (
            DEFAULT_MIN_EXPLORE_SCORE
        ),
    ) -> OutreachReportSummary:
        rows = (
            self.repository
            .list_report_rows(
                search_profile_name=(
                    search_profile_name
                ),
                min_score=0.0,
            )
        )

        priority: list[
            OutreachReportRow
        ] = []

        explore: list[
            OutreachReportRow
        ] = []

        history: list[
            OutreachReportRow
        ] = []

        for row in rows:
            if row.contacted:
                history.append(
                    row
                )
                continue

            if (
                row.best_contact_id is None
                or row.contact_type is None
                or row.contact_value is None
            ):
                continue

            contact_score = (
                _contact_score(
                    row
                )
            )

            if contact_score <= 0:
                continue

            if (
                row.score >= min_score
                and _has_priority_evidence(
                    row
                )
            ):
                priority.append(
                    row
                )
                continue

            if (
                row.score
                >= min_explore_score
            ):
                explore.append(
                    row
                )

        workbook = Workbook()

        active = workbook.active
        active.title = (
            "Priority Outreach"
        )

        _write_sheet(
            active,
            priority,
            bucket="PRIORITY",
        )

        explore_sheet = (
            workbook.create_sheet(
                "Explore"
            )
        )

        _write_sheet(
            explore_sheet,
            explore,
            bucket="EXPLORE",
        )

        history_sheet = (
            workbook.create_sheet(
                "History"
            )
        )

        _write_sheet(
            history_sheet,
            history,
            bucket="HISTORY",
        )

        output.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(
            output
        )

        return OutreachReportSummary(
            priority=len(
                priority
            ),
            explore=len(
                explore
            ),
            history=len(
                history
            ),
        )


HEADERS = (
    "Score",
    "Level",
    "Bucket",
    "Company",
    "Contact",
    "Contact Type",
    "Contact Quality",
    "Contact Score",
    "Website",
    "Careers",
    "Historical Match",
    "Current Match",
    "Current Relevant Jobs",
    "Manual Reference",
    "CESSI Source",
    "Country",
    "Remote Argentina",
    "Remote LATAM",
    "Company Type",
    "Target Priority",
    "Reasons",
    "Contact Source",
    "Outreach Status",
    "Outreach At",
    "Company ID",
    "Contact ID",
)


def _write_sheet(
    sheet,
    rows: list[
        OutreachReportRow
    ],
    *,
    bucket: str,
) -> None:
    sheet.append(
        HEADERS
    )

    for cell in sheet[1]:
        cell.font = Font(
            bold=True
        )

    for row in rows:
        contact_score = (
            _contact_score(
                row
            )
        )

        contact_quality = (
            _contact_quality(
                row
            )
        )

        sheet.append(
            (
                row.score,
                row.level,
                bucket,
                row.company_name,
                row.contact_value,
                row.contact_type,
                contact_quality,
                contact_score,
                row.website_url,
                row.careers_url,
                row.historical_max_match,
                row.current_max_match,
                row.current_relevant_jobs,
                _yes_no(
                    row.manual_reference
                ),
                _yes_no(
                    row.cessi_source
                ),
                row.country,
                _yes_no(
                    row.remote_argentina
                ),
                _yes_no(
                    row.remote_latam
                ),
                row.company_type,
                row.target_priority,
                "; ".join(
                    row.reasons
                ),
                row.contact_source_url,
                row.outreach_status,
                row.outreach_at,
                row.company_id,
                row.best_contact_id,
            )
        )

        current_row = (
            sheet.max_row
        )

        contact_cell = sheet.cell(
            row=current_row,
            column=5,
        )

        if row.contact_value:
            if row.contact_type in {
                "RECRUITING_EMAIL",
                "CAREERS_EMAIL",
                "GENERAL_EMAIL",
            }:
                contact_cell.hyperlink = (
                    "mailto:"
                    + row.contact_value
                )

            elif (
                row.contact_type
                == "GENERAL_APPLICATION_URL"
            ):
                contact_cell.hyperlink = (
                    row.contact_value
                )

        for column in (
            9,
            10,
            22,
        ):
            cell = sheet.cell(
                row=current_row,
                column=column,
            )

            if (
                isinstance(
                    cell.value,
                    str,
                )
                and cell.value.startswith(
                    (
                        "http://",
                        "https://",
                    )
                )
            ):
                cell.hyperlink = (
                    cell.value
                )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    widths = (
        10,
        12,
        12,
        30,
        34,
        24,
        18,
        14,
        34,
        34,
        16,
        14,
        20,
        18,
        14,
        18,
        18,
        14,
        18,
        18,
        56,
        34,
        18,
        24,
        12,
        12,
    )

    for index, width in enumerate(
        widths,
        start=1,
    ):
        sheet.column_dimensions[
            get_column_letter(
                index
            )
        ].width = width

    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def _has_priority_evidence(
    row: OutreachReportRow,
) -> bool:
    return (
        row.current_max_match
        is not None
        or row.historical_max_match
        is not None
        or row.manual_reference
    )


def _contact_score(
    row: OutreachReportRow,
) -> float:
    if (
        row.contact_type is None
        or row.contact_value is None
    ):
        return 0.0

    return (
        contact_quality_score_for(
            ContactType(
                row.contact_type
            ),
            row.contact_value,
        )
    )


def _contact_quality(
    row: OutreachReportRow,
) -> str:
    if (
        row.contact_type is None
        or row.contact_value is None
    ):
        return ""

    return (
        contact_quality_label_for(
            ContactType(
                row.contact_type
            ),
            row.contact_value,
        )
    )


def _yes_no(
    value: bool | None,
) -> str:
    if value is True:
        return "YES"

    if value is False:
        return "NO"

    return ""
